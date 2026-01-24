#!/usr/bin/env python3
"""
Update Surgery Metadata Script

Reads surgery information from PhysicalExamination.xlsx and appends it
to metadata.yaml files in the RM patient directories.

Usage:
    python scripts/update_surgery_metadata.py --dry-run   # Preview changes
    python scripts/update_surgery_metadata.py             # Execute updates
"""

import os
import re
import argparse
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import yaml
from openpyxl import load_workbook

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SurgeryMetadataUpdater:
    """Updates RM metadata files with surgery information from Excel."""

    # Column indices in PhysicalExamination.xlsx (0-indexed)
    COL_PID = 0
    COL_SURGERY1 = 77
    COL_SURGERY2 = 79

    def __init__(self,
                 rm_root: str = "/mnt/blue8T/CP/RM",
                 excel_path: str = "/mnt/blue8T/CP/PhysicalExamination.xlsx",
                 alias_path: str = None,
                 dry_run: bool = False):
        self.rm_root = Path(rm_root)
        self.excel_path = Path(excel_path)
        self.dry_run = dry_run

        # Load alias table
        if alias_path is None:
            alias_path = self.rm_root / "operations_alias.yaml"
        self.aliases = self._load_aliases(alias_path)

    def _load_aliases(self, path: Path) -> Dict[str, str]:
        """Load operation name aliases from YAML file."""
        if not path.exists():
            logger.warning(f"Alias file not found: {path}")
            return {}

        with open(path) as f:
            data = yaml.safe_load(f)

        return data.get('aliases', {})

    def resolve_procedure_name(self, raw_name: str) -> str:
        """
        Resolve procedure name through alias table.

        1. Strip whitespace
        2. Check if raw_name exists in aliases dict
        3. If yes, return canonical name
        4. If no, replace spaces with underscores and return
        """
        name = raw_name.strip()

        # Remove parenthetical notes like "(plate)"
        name = re.sub(r'\s*\([^)]*\)\s*$', '', name)
        name = name.strip()

        if name in self.aliases:
            return self.aliases[name]

        return name.replace(' ', '_')

    def parse_surgery_string(self, surgery_str: str) -> List[str]:
        """
        Parse surgery string and return list of procedures with side suffixes.

        Examples:
            "FDO, TAL, TPST, PF, Rt." → [FDO_Rt, TAL_Rt, TPST_Rt, PF_Rt]
            "FDO, DHL, bilateral + TAL, Rt." → [FDO_Rt, FDO_Lt, DHL_Rt, DHL_Lt, TAL_Rt]
        """
        if not surgery_str or not surgery_str.strip():
            return []

        # Remove parenthetical notes like "(plate)" anywhere in the string
        cleaned_str = re.sub(r'\s*\([^)]*\)', '', surgery_str)

        procedures = []

        # Split by ' + ' to separate side-specific groups
        # Also handle cases where + is adjacent to text
        segments = re.split(r'\s*\+\s*', cleaned_str)

        for segment in segments:
            segment = segment.strip()
            if not segment:
                continue

            # Determine side indicator
            side = self._extract_side(segment)

            # Remove side indicator from segment to get procedure list
            cleaned_segment = self._remove_side_indicator(segment)

            # Split by comma to get individual procedures
            proc_names = [p.strip() for p in cleaned_segment.split(',') if p.strip()]

            for proc_name in proc_names:
                # Skip empty entries
                if not proc_name:
                    continue

                # Skip if it's just a side indicator that wasn't cleaned
                if proc_name.lower() in ['rt', 'lt', 'rt.', 'lt.', 'bilateral', 'b/l', 'b/l.']:
                    continue

                # Resolve alias
                canonical_name = self.resolve_procedure_name(proc_name)

                # Skip if canonical name is empty after processing
                if not canonical_name:
                    continue

                # Apply side suffix
                if side == 'bilateral':
                    procedures.append(f"{canonical_name}_Rt")
                    procedures.append(f"{canonical_name}_Lt")
                elif side in ['Rt', 'Lt']:
                    procedures.append(f"{canonical_name}_{side}")
                else:
                    # No side specified, add without suffix
                    procedures.append(canonical_name)

        return procedures

    def _extract_side(self, segment: str) -> Optional[str]:
        """Extract side indicator from segment."""
        segment_lower = segment.lower().strip()

        # Check for bilateral indicators at the end
        # Match patterns like: B/L, B/L., bilateral, b./b.l.
        if re.search(r'(bilateral|b/l\.?|b\./b\.l\.)\s*$', segment_lower):
            return 'bilateral'

        # Check for right side (at end of string)
        if re.search(r'\brt\.?\s*$', segment_lower):
            return 'Rt'

        # Check for left side (at end of string)
        if re.search(r'\blt\.?\s*$', segment_lower):
            return 'Lt'

        return None

    def _remove_side_indicator(self, segment: str) -> str:
        """Remove side indicator from segment."""
        # Remove bilateral indicators (including B/L. with period)
        result = re.sub(r',?\s*(bilateral|b/l\.?|b\./b\.l\.)\s*$', '', segment, flags=re.IGNORECASE)

        # Remove Rt. or Lt. at the end (with optional period and spaces)
        result = re.sub(r',?\s*(rt|lt)\.?\s*$', '', result, flags=re.IGNORECASE)

        return result.strip()

    def read_excel_data(self) -> Dict[str, Tuple[Optional[str], Optional[str]]]:
        """
        Read surgery data from Excel file.

        Returns:
            Dict mapping PID to (surgery1, surgery2) strings
        """
        logger.info(f"Reading Excel file: {self.excel_path}")

        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb.active

        data = {}

        # Data starts from row 3 (row 1 and 2 are headers)
        for row in ws.iter_rows(min_row=3):
            pid_cell = row[self.COL_PID]

            # Skip rows without patient ID
            if not pid_cell.value:
                continue

            pid = str(pid_cell.value).strip()

            # Skip header rows
            if pid == '환자번호' or not pid.isdigit():
                continue

            surgery1 = row[self.COL_SURGERY1].value if len(row) > self.COL_SURGERY1 else None
            surgery2 = row[self.COL_SURGERY2].value if len(row) > self.COL_SURGERY2 else None

            # Only include patients with surgery data
            if surgery1 or surgery2:
                data[pid] = (surgery1, surgery2)

        wb.close()

        logger.info(f"Found {len(data)} patients with surgery data")
        return data

    def get_rm_patients(self) -> List[str]:
        """Get list of patient IDs in RM directory."""
        patients = []
        for item in self.rm_root.iterdir():
            if item.is_dir() and item.name.isdigit():
                patients.append(item.name)
        return sorted(patients)

    def update_metadata_file(self, metadata_path: Path, surgery_list: List[str]) -> bool:
        """
        Update a metadata.yaml file with surgery information.

        Returns:
            True if update was successful (or would be in dry-run)
        """
        if not metadata_path.exists():
            logger.warning(f"Metadata file not found: {metadata_path}")
            return False

        # Read existing metadata
        with open(metadata_path) as f:
            metadata = yaml.safe_load(f) or {}

        # Add surgery list
        metadata['surgery'] = surgery_list

        if self.dry_run:
            logger.info(f"  [DRY-RUN] Would update: {metadata_path}")
            logger.info(f"    Surgery: {surgery_list}")
            return True

        # Write updated metadata
        with open(metadata_path, 'w') as f:
            yaml.dump(metadata, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

        logger.debug(f"  Updated: {metadata_path}")
        return True

    def create_op2_metadata(self, op2_dir: Path, surgery_list: List[str]) -> bool:
        """
        Create op2 directory and minimal metadata.yaml with surgery info.

        Returns:
            True if creation was successful (or would be in dry-run)
        """
        if self.dry_run:
            logger.info(f"  [DRY-RUN] Would create: {op2_dir}")
            logger.info(f"  [DRY-RUN] Would create: {op2_dir / 'metadata.yaml'}")
            logger.info(f"    Surgery: {surgery_list}")
            return True

        # Create op2 directory
        op2_dir.mkdir(parents=True, exist_ok=True)

        # Create minimal metadata.yaml
        metadata = {'surgery': surgery_list}

        metadata_path = op2_dir / 'metadata.yaml'
        with open(metadata_path, 'w') as f:
            yaml.dump(metadata, f, allow_unicode=True, default_flow_style=False)

        logger.debug(f"  Created: {metadata_path}")
        return True

    def process_patient(self, pid: str, surgery1: Optional[str], surgery2: Optional[str]) -> bool:
        """
        Process surgery data for a single patient.

        Returns:
            True if processing was successful
        """
        patient_dir = self.rm_root / pid

        if not patient_dir.exists():
            logger.warning(f"Patient directory not found: {patient_dir}")
            return False

        success = True

        # Process surgery1 → op1/metadata.yaml
        if surgery1:
            surgery1_list = self.parse_surgery_string(surgery1)
            if surgery1_list:
                op1_metadata = patient_dir / "op1" / "metadata.yaml"
                if op1_metadata.exists():
                    if not self.update_metadata_file(op1_metadata, surgery1_list):
                        success = False
                else:
                    logger.warning(f"  op1/metadata.yaml not found for {pid}")

        # Process surgery2 → op2/metadata.yaml (create if needed)
        if surgery2:
            surgery2_list = self.parse_surgery_string(surgery2)
            if surgery2_list:
                op2_dir = patient_dir / "op2"
                if not self.create_op2_metadata(op2_dir, surgery2_list):
                    success = False

        return success

    def run(self, patient_filter: Optional[str] = None) -> Tuple[int, int]:
        """
        Run the update process.

        Args:
            patient_filter: Optional single patient ID to process

        Returns:
            Tuple of (success_count, fail_count)
        """
        # Read Excel data
        excel_data = self.read_excel_data()

        # Get RM patients
        rm_patients = set(self.get_rm_patients())

        # Determine which patients to process
        if patient_filter:
            patients_to_process = [patient_filter] if patient_filter in excel_data else []
        else:
            # Only process patients that exist in both Excel and RM directory
            patients_to_process = sorted(set(excel_data.keys()) & rm_patients)

        logger.info(f"Processing {len(patients_to_process)} patients...")

        success_count = 0
        fail_count = 0

        for pid in patients_to_process:
            surgery1, surgery2 = excel_data[pid]

            logger.info(f"Processing patient {pid}...")

            if self.process_patient(pid, surgery1, surgery2):
                success_count += 1
            else:
                fail_count += 1

        return success_count, fail_count


def main():
    parser = argparse.ArgumentParser(
        description='Update RM metadata files with surgery information'
    )
    parser.add_argument(
        '--dry-run', '-n',
        action='store_true',
        help='Preview changes without writing files'
    )
    parser.add_argument(
        '--patient', '-p',
        type=str,
        help='Process only a specific patient ID'
    )
    parser.add_argument(
        '--rm-root',
        type=str,
        default='/mnt/blue8T/CP/RM',
        help='RM root directory'
    )
    parser.add_argument(
        '--excel-path',
        type=str,
        default='/mnt/blue8T/CP/PhysicalExamination.xlsx',
        help='Path to PhysicalExamination.xlsx'
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Verbose output'
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    updater = SurgeryMetadataUpdater(
        rm_root=args.rm_root,
        excel_path=args.excel_path,
        dry_run=args.dry_run
    )

    success_count, fail_count = updater.run(patient_filter=args.patient)

    print(f"\nSummary: {success_count} succeeded, {fail_count} failed")

    if args.dry_run:
        print("\n[DRY-RUN] No files were modified")


if __name__ == '__main__':
    main()
